import tkinter as tk
import os
import sqlite3
from tkinter import ttk
from tkinter import messagebox
import openpyxl as xl
from openpyxl import styles
from openpyxl.utils import get_column_letter
from datetime import datetime


class App(tk.Tk):
    rate = 0.0
    current_job = 0

    # Get the time, rounded to the nearest 15 minutes
    def get_time(self):
        return round(datetime.now().timestamp() / 900.0) * 900

    # Perform a query
    def do_query(self, iQuery, iArgs):
        database = os.environ['HOME'] + '/wfh_log.sqlite'
        try:
            connection = sqlite3.connect(database)
        except Exception as e:
            print("Error: {}".format(str(e)))
            return []
        try:
            cursor = connection.cursor()
        except Exception as e:
            print("Error: {}".format(str(e)))
            return []
        try:
            cursor.execute(iQuery, iArgs)
        except Exception as e:
            print("Query is: {}".format(iQuery))
            print("Args are: {}".format(iArgs))
            print("Error: {}".format(str(e)))
            return []
        connection.commit()
        rows = cursor.fetchall()
        # debug: print the first row to console
        # print(rows[0])
        connection.close()
        return rows
        # end of do_query

    def show_log(self, tv):
        # Clear the treeview first
        for row in tv.get_children():
            tv.delete(row)
        query = """
            SELECT rate FROM wfh_rate ORDER by start_date DESC LIMIT 1
        """
        try:
            rows = self.do_query(query, ())
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        self.rate = rows[0][0]
        query = """
            SELECT * FROM wfh_log
            WHERE id > (SELECT max(id) FROM wfh_log) - 10
            ORDER BY start_time LIMIT 10
        """
        try:
            rows = self.do_query(query, ())
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        for row in rows:
            value_list = []
            try:
                dStart = datetime.fromtimestamp(row[1])
            except Exception as e:
                print("Error: {}".format(str(e)))
                dStart = datetime.now()
            try:
                dFinish = datetime.fromtimestamp(row[2])
            except Exception as e:
                print("Error: {}".format(str(e)))
                dFinish = datetime.now()
            dHours = (row[2] - row[1]) / 3600
            value_list.append(format(dStart, "%Y-%m-%d"))
            value_list.append(format(dStart, "%a"))
            value_list.append(format(dStart, "%H:%M"))
            value_list.append(format(dFinish, "%H:%M"))
            value_list.append(format(dHours, ".4f"))
            value_list.append(format(self.rate * dHours, ".4f"))
            value_list.append(row[3])
            tv.insert(
                '',
                'end',
                text="1",
                values=value_list
            )

    def start_log(self, tv, start, reason):
        # If there is a job in progress, don't start a new one
        if self.current_job != 0:
            messagebox.showerror(
                "Error",
                "There is already a job in progress"
            )
            return
        if reason == "":
            messagebox.showerror(
                "Error",
                "Please enter a reason"
            )
        # Get the id of the last row
        query = """
            SELECT id FROM wfh_log ORDER BY id DESC LIMIT 1
        """
        try:
            rows = self.do_query(query, ())
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        self.current_job = rows[0][0] + 1
        # Insert current Date, Day, time, and reason into spreadsheet
        # Initialise the finish_time to start plus 8 hours
        finish = start + 3600 * 8
        query = """
            INSERT INTO wfh_log (id, start_time, finish_time, reason)
            VALUES (?, ?, ?, ?)
        """
        try:
            self.do_query(query, (self.current_job, start, finish, reason))
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        self.show_log(tv)

    def finish_log(self, tv, finish, reason):
        # If there is no job in progress, don't finish
        if self.current_job == 0:
            messagebox.showerror(
                "Error",
                "There is no job in progress"
            )
            return
        # Put the current time and reason, in case modified, into db
        query = """
            UPDATE wfh_log SET finish_time = ?, reason = ? WHERE id = ?
        """
        try:
            self.do_query(query, (finish, reason, self.current_job))
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        # Set the current job to 0
        self.current_job = 0
        self.show_log(tv)

    def export_log(self):
        # Get all the records from the log
        query = """
            SELECT * FROM wfh_log
        """
        try:
            rows = self.do_query(query, ())
        except Exception as e:
            print("Error: {}".format(str(e)))
            return

        # Export the data to a spreadsheeet
        dateStr = datetime.now().strftime("%Y%m%d%H%M")
        xlFile = os.environ['HOME'] + '/wfh_log_' + dateStr + '.xlsx'
        book = xl.Workbook()
        sheet = book.active
        rowNumber = 0
        sheet.cell(1, 1).value = 'Date'
        sheet.cell(1, 2).value = 'Day'
        sheet.cell(1, 3).value = 'Start'
        sheet.cell(1, 4).value = 'Finish'
        sheet.cell(1, 5).value = 'Hours'
        sheet.cell(1, 6).value = 'Amt'
        sheet.cell(1, 7).value = 'Reason'
        column_widths = [15, 10, 10, 10, 10, 10, 40]
        hFont = styles.Font(bold=True)
        for cells in sheet['A1:G1']:
            for cell in cells:
                cell.font = hFont
        hAlign = styles.Alignment(
            horizontal='center',
            vertical='bottom'
        )
        for cells in sheet['A1:G1']:
            for cell in cells:
                cell.alignment = hAlign
        for row in rows:
            start_time = datetime.fromtimestamp(row[1])
            finish_time = datetime.fromtimestamp(row[2])
            sheet.cell(rowNumber + 2, 1).value = start_time
            sheet.cell(rowNumber + 2, 1).number_format = 'yyyy-mm-dd'
            sheet.cell(rowNumber + 2, 2).value = start_time
            sheet.cell(rowNumber + 2, 2).number_format = 'ddd'
            sheet.cell(rowNumber + 2, 3).value = start_time
            sheet.cell(rowNumber + 2, 3).number_format = 'hh:mm'
            sheet.cell(rowNumber + 2, 4).value = finish_time
            sheet.cell(rowNumber + 2, 4).number_format = 'hh:mm'
            sheet.cell(rowNumber + 2, 5).value = \
                (finish_time - start_time) / 3600.0
            sheet.cell(rowNumber + 2, 5).number_format = '0.0000'
            sheet.cell(rowNumber + 2, 6).value = \
                (finish_time - start_time) / 3600.0 * self.rate
            sheet.cell(rowNumber + 2, 6).number_format = '0.0000'
            sheet.cell(rowNumber + 2, 7).value = row[3]
            rowNumber += 1
        for rowNumber in range(1, sheet.max_row):
            for colNumber in range(1, sheet.max_column):
                if len(sheet.cell(rowNumber, colNumber).number_format) > \
                  column_widths[colNumber - 1]:
                    column_widths[colNumber - 1] = \
                        len(sheet.cell(rowNumber, colNumber).number_format)
        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i + 1)].width = \
                column_width
        book.save(xlFile)
        book.close()
        messagebox.showinfo(
            'Export',
            'Log Exported to {}'.format(xlFile)
        )

    def delete_log(self, tv):
        # Find the last job in the log
        query = """
            SELECT id FROM wfh_log ORDER BY id DESC LIMIT 1
        """
        try:
            rows = self.do_query(query, ())
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        self.current_job = rows[0][0]
        query = """
            DELETE FROM wfh_log WHERE id = ?
        """
        try:
            self.do_query(query, (self.current_job,))
        except Exception as e:
            print("Error: {}".format(str(e)))
            return
        self.current_job = 0
        self.show_log(tv)

    def __init__(self):
        super().__init__()

        # root window
        self.title('Work From Home Log')
        # self.geometry('720x600')
        self.style = ttk.Style(self)
        self.style.theme_use('clam')

        # Reason Label
        label = ttk.Label(self, text='Reason:')
        label.grid(column=0, row=0, padx=10, pady=10,  sticky='w')
        # Reason
        reason = ttk.Entry(self)
        reason.grid(column=1, row=0, padx=10, pady=10,  sticky='w')
        # View of the last 10 rows of the spreadsheet
        wfh_log = ttk.Treeview(
            self,
            columns=(
                'Date',
                'DOW',
                'Start',
                'Finish',
                'Hours',
                'Amt',
                'Reason'
            ),
            show='headings',
            height=10,
            selectmode='browse'
        )
        wfh_log.grid(
            column=0,
            row=1,
            rowspan=10,
            columnspan=7,
            sticky='nsew'
        )
        wfh_log.column('#1', width=80, anchor='w')
        wfh_log.heading('#1', text='Date')
        wfh_log.column('#2', width=50, anchor='w')
        wfh_log.heading('#2', text='DOW')
        wfh_log.column('#3', width=50, anchor='e')
        wfh_log.heading('#3', text='Start')
        wfh_log.column('#4', width=50, anchor='e')
        wfh_log.heading('#4', text='Finish')
        wfh_log.column('#5', width=50, anchor='e')
        wfh_log.heading('#5', text='Hours')
        wfh_log.column('#6', width=80, anchor='e')
        wfh_log.heading('#6', text='Amt')
        wfh_log.column('#7', width=250, anchor='w')
        wfh_log.heading('#7', text='Reason')
        # Start Button
        btnStart = ttk.Button(
            self,
            text='Start',
            command=lambda: self.start_log(
                wfh_log,
                self.get_time(),
                reason.get()
            )
        )
        btnStart.grid(column=2, row=0, padx=10, pady=10,  sticky='w')
        # Stop Button
        btnStop = ttk.Button(
            self,
            text='Stop',
            command=lambda: self.finish_log(
                wfh_log,
                self.get_time(),
                reason.get()
            )
        )
        btnStop.grid(column=3, row=0, padx=10, pady=10,  sticky='w')
        # Delete Button
        btnDelete = ttk.Button(
            self,
            text='Delete',
            command=lambda: self.delete_log(wfh_log)
        )
        btnDelete.grid(column=4, row=0, padx=10, pady=10,  sticky='w')
        # Export Button
        btnExport = ttk.Button(
            self,
            text='Export',
            command=lambda: self.export_log()
        )
        btnExport.grid(column=5, row=0, padx=10, pady=10,  sticky='w')
        # Quit Button
        btnQuit = ttk.Button(
            self,
            text='Quit',
            command=self.destroy
        )
        btnQuit.grid(column=6, row=0, padx=10, pady=10,  sticky='w')
        self.show_log(wfh_log)
        self.mainloop()


if __name__ == "__main__":
    app = App()
    app.mainloop()
